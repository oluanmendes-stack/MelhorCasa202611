#!/usr/bin/env python3
"""
Flet-based GUI for real estate scraping
Mirrors the functionality of the React frontend but runs as a standalone desktop application
"""

import flet as ft
import subprocess
import json
import threading
import sys
import os
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field
from datetime import datetime
import math


@dataclass
class Property:
    id: str
    nome: str
    imagem: str
    valor: str
    m2: str
    localizacao: str
    link: str
    quartos: str
    garagem: str
    banhos: str = ""
    site: str = ""
    tags: List[str] = field(default_factory=list)
    latitude: Optional[float] = None
    longitude: Optional[float] = None

    def __hash__(self):
        return hash(self.link)

    def __eq__(self, other):
        if isinstance(other, Property):
            return self.link == other.link
        return False


def parse_int(s: str) -> int:
    """Parse integer from string with non-digit characters"""
    try:
        return int("".join([c for c in str(s) if c.isdigit()]))
    except Exception:
        return 0


def sanitize_image(url: str) -> str:
    """Sanitize image URL"""
    if not url:
        return ""
    if url.startswith("http:"):
        return url.replace("http:", "https:", 1)
    if url.startswith("//"):
        return "https:" + url
    return url


class ScraperGUI:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = "Melhor Casa - Web Scraper"
        self.page.window_width = 1400
        self.page.window_height = 900
        self.page.theme_mode = ft.ThemeMode.LIGHT
        
        self.scraping = False
        self.properties: List[Property] = []
        self.liked_properties: List[Property] = []
        self.disliked_properties: List[Property] = []
        self.user_location: Optional[Dict[str, Any]] = None
        
        # Current view
        self.current_view = "busca"  # busca, curtidas, rejeitadas, ranking
        
        # Filter and sorting states
        self.filters = {
            "valorMin": "",
            "valorMax": "250000",
            "m2Min": 0,
            "m2Max": float('inf'),
            "quartos": ["all"],
            "vagas": ["all"],
            "banhos": [],
            "distanciaMax": float('inf'),
        }
        
        self.sort_option = {"field": "valor", "direction": "asc"}
        self.status_filters = {"naPlanta": True, "emConstrucao": True, "leilao": True}
        
        # Preferences for ranking
        self.pref_tamanho_priority = 1
        self.pref_quartos_priority = 2
        self.pref_banheiros_priority = 3
        self.pref_distancia_priority = 4
        
        self.pref_tamanho_value = 80
        self.pref_quartos_value = 2
        self.pref_banheiros_value = 1
        self.pref_distancia_value = 10
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the main UI layout with tabs"""
        # Header
        header = ft.Container(
            content=ft.Row(
                controls=[
                    ft.Icon(ft.Icons.HOME, size=32, color="#2563eb"),
                    ft.Column(
                        controls=[
                            ft.Text("Melhor Casa", size=24, weight="bold"),
                            ft.Text("Ferramenta de coleta de imóveis", size=12, color="gray"),
                        ],
                        spacing=0
                    ),
                    ft.Container(expand=True),
                    ft.IconButton(
                        icon=ft.Icons.SETTINGS,
                        on_click=lambda e: self.show_settings_dialog(),
                        tooltip="Configurações"
                    ),
                ],
                spacing=16,
                vertical_alignment="center",
            ),
            padding=16,
            bgcolor="#f9fafb",
            border_radius=8,
        )
        
        # Navigation tabs
        self.tabs = ft.Tabs(
            selected_index=0,
            tabs=[
                ft.Tab(
                    text="🔍 Busca",
                    content=ft.Column(
                        controls=[
                            # Search and filter controls
                            self.build_search_controls(),
                            self.build_filter_controls(),
                            ft.Divider(),
                            # Properties header
                            ft.Row(
                                controls=[
                                    ft.Text("Imóveis Encontrados", size=16, weight="bold"),
                                    ft.Container(expand=True),
                                    ft.IconButton(
                                        icon=ft.Icons.FAVORITE,
                                        tooltip="Ver favoritos",
                                        on_click=lambda e: self.switch_tab(1),
                                    ),
                                ],
                                spacing=8,
                            ),
                            # Properties view expands to fill remaining space
                            self.build_properties_view(),
                        ],
                        expand=True,
                        spacing=8,
                        padding=16,
                        scroll="auto",
                    ),
                    icon="search",
                ),
                ft.Tab(
                    text="❤️ Curtidas",
                    content=ft.Container(
                        content=ft.Column(
                            controls=[
                                ft.Row(
                                    controls=[
                                        ft.Text("Imóveis Favoritos", size=16, weight="bold"),
                                        ft.Container(expand=True),
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE,
                                            tooltip="Limpar favoritos",
                                            on_click=self.clear_liked_properties,
                                        ),
                                    ],
                                    spacing=8,
                                ),
                                self.build_liked_view(),
                            ],
                            expand=True,
                            spacing=12,
                        ),
                        padding=16,
                        expand=True,
                    ),
                    icon="favorite",
                ),
                ft.Tab(
                    text="👎 Rejeitadas",
                    content=ft.Container(
                        content=ft.Column(
                            controls=[
                                ft.Row(
                                    controls=[
                                        ft.Text("Imóveis Rejeitados", size=16, weight="bold"),
                                        ft.Container(expand=True),
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE,
                                            tooltip="Limpar rejeitados",
                                            on_click=self.clear_disliked_properties,
                                        ),
                                    ],
                                    spacing=8,
                                ),
                                self.build_disliked_view(),
                            ],
                            expand=True,
                            spacing=12,
                        ),
                        padding=16,
                        expand=True,
                    ),
                    icon="thumb_down",
                ),
                ft.Tab(
                    text="📊 Ranking",
                    content=ft.Container(
                        content=ft.Column(
                            controls=[
                                ft.Row(
                                    controls=[
                                        ft.Text("Ranking de Imóveis", size=16, weight="bold"),
                                        ft.Container(expand=True),
                                        ft.IconButton(
                                            icon=ft.Icons.REFRESH,
                                            tooltip="Atualizar ranking",
                                            on_click=self.refresh_ranking,
                                        ),
                                    ],
                                    spacing=8,
                                ),
                                self.build_ranking_view(),
                            ],
                            expand=True,
                            spacing=12,
                        ),
                        padding=16,
                        expand=True,
                    ),
                    icon="sort",
                ),
            ],
            expand=True,
        )
        
        # Status bar
        self.status_text = ft.Text("Pronto", size=12, color="gray")
        self.status_indicator = ft.Container(
            width=12,
            height=12,
            border_radius=50,
            bgcolor="gray",
        )
        
        self.status_stats = ft.Text(f"Total: 0 | Favoritos: 0 | Rejeitados: 0", size=12, color="gray")
        
        status_bar = ft.Container(
            content=ft.Row(
                controls=[
                    self.status_indicator,
                    self.status_text,
                    ft.Container(expand=True),
                    self.status_stats,
                ],
                spacing=16,
                alignment="center",
            ),
            padding=8,
            bgcolor="#f0f4f8",
            border_radius=8,
        )
        
        # Main layout
        main_content = ft.Container(
            content=ft.Column(
                controls=[
                    header,
                    ft.Divider(),
                    self.tabs,
                    ft.Divider(),
                    status_bar,
                ],
                expand=True,
                spacing=8,
            ),
            padding=12,
            expand=True,
        )
        
        self.page.add(main_content)
    
    def build_properties_view(self) -> ft.Container:
        """Build the properties display container"""
        self.properties_view = ft.Column(
            scroll="auto",
            expand=True,
            spacing=12,
        )
        return ft.Container(
            content=self.properties_view,
            expand=True,
        )
    
    def build_liked_view(self) -> ft.Container:
        """Build the liked properties view"""
        self.liked_view = ft.Column(
            scroll="auto",
            expand=True,
            spacing=12,
        )
        return ft.Container(
            content=self.liked_view,
            expand=True,
        )
    
    def build_disliked_view(self) -> ft.Container:
        """Build the disliked properties view"""
        self.disliked_view = ft.Column(
            scroll="auto",
            expand=True,
            spacing=12,
        )
        return ft.Container(
            content=self.disliked_view,
            expand=True,
        )
    
    def build_ranking_view(self) -> ft.Container:
        """Build the ranking view"""
        self.ranking_view = ft.Column(
            scroll="auto",
            expand=True,
            spacing=12,
        )
        return ft.Container(
            content=self.ranking_view,
            expand=True,
        )
    
    def switch_tab(self, index: int):
        """Switch to a different tab"""
        self.tabs.selected_index = index
        self.page.update()
    
    def build_search_controls(self) -> ft.Card:
        """Build the search/scraping controls section"""
        # Site checkboxes
        sites = [
            ("Netimóveis", "netimoveis"),
            ("Casa Mineira", "casamineira"),
            ("Imóvel Web", "imovelweb"),
            ("Zap Imóveis", "zapimoveis"),
            ("Viva Real", "vivareal"),
            ("OLX", "olx"),
            ("Quinto Andar", "quintoandar"),
            ("Loft", "loft"),
            ("Chaves na Mão", "chavesnamao"),
        ]
        
        self.site_checks = {}
        site_controls = []
        
        for label, key in sites:
            cb = ft.Checkbox(label=label, value=False)
            self.site_checks[key] = cb
            site_controls.append(cb)
        
        sites_column = ft.Column(
            controls=[
                ft.Text("Selecione os sites", weight="bold", size=12),
                ft.Row(
                    controls=site_controls[:3],
                    wrap=True,
                    spacing=8,
                ),
                ft.Row(
                    controls=site_controls[3:6],
                    wrap=True,
                    spacing=8,
                ),
                ft.Row(
                    controls=site_controls[6:],
                    wrap=True,
                    spacing=8,
                ),
            ],
            spacing=8,
        )
        
        # Location inputs
        self.cidade_input = ft.TextField(
            label="Cidade",
            value="Belo Horizonte",
            width=150,
        )
        
        self.bairro_input = ft.TextField(
            label="Bairro (Opcional)",
            value="",
            width=150,
        )
        
        self.tipo_select = ft.Dropdown(
            label="Tipo de Imóvel",
            options=[
                ft.dropdown.Option("indiferente", "Indiferente"),
                ft.dropdown.Option("apartamentos", "Apartamentos"),
                ft.dropdown.Option("casas", "Casas"),
                ft.dropdown.Option("garagens", "Garagem"),
                ft.dropdown.Option("estacionamento", "Estacionamento"),
            ],
            value="apartamentos",
            width=150,
        )
        
        location_row = ft.Row(
            controls=[
                self.cidade_input,
                self.bairro_input,
                self.tipo_select,
            ],
            spacing=8,
            wrap=True,
        )
        
        # Control buttons
        self.start_button = ft.ElevatedButton(
            text="Iniciar Scraping",
            icon=ft.Icons.PLAY_CIRCLE_FILLED,
            on_click=self.on_start_scraping,
            width=150,
        )

        self.stop_button = ft.ElevatedButton(
            text="Parar",
            icon=ft.Icons.STOP_CIRCLE,
            on_click=self.on_stop_scraping,
            disabled=True,
            width=150,
        )

        self.export_button = ft.ElevatedButton(
            text="Exportar Excel",
            icon=ft.Icons.GET_APP,
            on_click=self.on_export_excel,
            width=150,
        )
        
        buttons_row = ft.Row(
            controls=[
                self.start_button,
                self.stop_button,
                self.export_button,
            ],
            spacing=8,
            wrap=True,
        )
        
        return ft.Card(
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Text("Procura", size=14, weight="bold"),
                        sites_column,
                        location_row,
                        buttons_row,
                    ],
                    spacing=12,
                ),
                padding=16,
            ),
            margin=0,
        )
    
    def build_filter_controls(self) -> ft.Card:
        """Build the filter controls section"""
        self.valormin_input = ft.TextField(
            label="Valor Mínimo (R$)",
            value="",
            width=120,
        )
        
        self.valormax_input = ft.TextField(
            label="Valor Máximo (R$)",
            value="250000",
            width=120,
        )
        
        self.area_min_input = ft.TextField(
            label="Área Mínima (m²)",
            value="",
            width=120,
        )
        
        self.area_max_input = ft.TextField(
            label="Área Máxima (m²)",
            value="",
            width=120,
        )
        
        self.quartos_input = ft.TextField(
            label="Quartos (Ex: 2,3,4+)",
            value="",
            width=120,
        )
        
        self.vagas_input = ft.TextField(
            label="Vagas (Ex: 1,2,3+)",
            value="",
            width=120,
        )
        
        self.banhos_input = ft.TextField(
            label="Banheiros (Ex: 1,2,3+)",
            value="",
            width=120,
        )
        
        # Sort options
        self.sort_field = ft.Dropdown(
            label="Ordenar por",
            options=[
                ft.dropdown.Option("valor", "Valor"),
                ft.dropdown.Option("tamanho", "Tamanho (m²)"),
                ft.dropdown.Option("distancia", "Distância"),
            ],
            value="valor",
            width=150,
            on_change=lambda e: self.on_sort_change(),
        )
        
        self.sort_direction = ft.Dropdown(
            label="Direção",
            options=[
                ft.dropdown.Option("asc", "Crescente"),
                ft.dropdown.Option("desc", "Decrescente"),
            ],
            value="asc",
            width=120,
            on_change=lambda e: self.on_sort_change(),
        )
        
        price_row = ft.Row(
            controls=[
                self.valormin_input,
                self.valormax_input,
            ],
            spacing=8,
            wrap=True,
        )
        
        area_row = ft.Row(
            controls=[
                self.area_min_input,
                self.area_max_input,
            ],
            spacing=8,
            wrap=True,
        )
        
        features_row = ft.Row(
            controls=[
                self.quartos_input,
                self.vagas_input,
                self.banhos_input,
            ],
            spacing=8,
            wrap=True,
        )
        
        sort_row = ft.Row(
            controls=[
                self.sort_field,
                self.sort_direction,
            ],
            spacing=8,
            wrap=True,
        )
        
        return ft.Card(
            content=ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Text("Filtros", size=14, weight="bold"),
                        price_row,
                        area_row,
                        features_row,
                        ft.Divider(height=8),
                        ft.Text("Ordenação", size=12, weight="bold"),
                        sort_row,
                    ],
                    spacing=12,
                ),
                padding=16,
            ),
            margin=0,
        )
    
    def on_sort_change(self):
        """Handle sort option changes"""
        self.sort_option["field"] = self.sort_field.value
        self.sort_option["direction"] = self.sort_direction.value
        self.refresh_properties_display()
    
    def on_start_scraping(self, e):
        """Start the scraping process"""
        if self.scraping:
            self.show_message("Scraping já em progresso", "warning")
            return
        
        # Check if at least one site is selected
        selected_sites = {k: v.value for k, v in self.site_checks.items()}
        if not any(selected_sites.values()):
            self.show_message("Selecione pelo menos um site", "error")
            return
        
        self.scraping = True
        self.start_button.disabled = True
        self.stop_button.disabled = False
        self.properties_view.controls.clear()
        self.page.update()
        
        # Run scraping in a separate thread to avoid blocking UI
        thread = threading.Thread(target=self.run_scraping, args=(selected_sites,))
        thread.daemon = True
        thread.start()
    
    def on_stop_scraping(self, e):
        """Stop the scraping process"""
        self.scraping = False
        self.start_button.disabled = False
        self.stop_button.disabled = True
        self.update_status("Scraping pausado", "warning")
    
    def on_export_excel(self, e):
        """Export properties to Excel file"""
        if not self.properties:
            self.show_message("Nenhum imóvel para exportar", "error")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.show_message("Instale 'openpyxl' para exportar Excel: pip install openpyxl", "error")
            return
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Imóveis"
            
            # Headers
            headers = ["Nome", "Valor", "M²", "Localização", "Quartos", "Garagem", "Banheiros", "Site", "Link", "Tags"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data
            for prop in self.properties:
                ws.append([
                    prop.nome,
                    prop.valor,
                    prop.m2,
                    prop.localizacao,
                    prop.quartos,
                    prop.garagem,
                    prop.banhos,
                    prop.site,
                    prop.link,
                    ", ".join(prop.tags) if prop.tags else "",
                ])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 20
            
            # Save file
            filename = f"imoveis_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            self.show_message(f"Arquivo salvo: {filename}", "success")
        except Exception as ex:
            self.show_message(f"Erro ao exportar: {str(ex)}", "error")
    
    def run_scraping(self, selected_sites: Dict[str, bool]):
        """Run the actual scraping in a background thread"""
        try:
            self.update_status("Iniciando scraping...", "info")
            
            # Build command arguments
            script_path = os.path.join(os.path.dirname(__file__), "scraper.py")
            args = [
                sys.executable,
                script_path,
                "--output", "json"
            ]
            
            # Add selected sites
            for site, selected in selected_sites.items():
                if selected:
                    args.append(f"--{site}")
            
            # Add filters
            if self.valormin_input.value:
                args.extend(["--valorMin", self.valormin_input.value])
            if self.valormax_input.value:
                args.extend(["--valorMax", self.valormax_input.value])
            if self.area_min_input.value:
                args.extend(["--areaMin", self.area_min_input.value])
            if self.area_max_input.value:
                args.extend(["--areaMax", self.area_max_input.value])
            if self.quartos_input.value:
                args.extend(["--quartos", self.quartos_input.value])
            if self.vagas_input.value:
                args.extend(["--vagas", self.vagas_input.value])
            if self.banhos_input.value:
                args.extend(["--banhos", self.banhos_input.value])
            
            # Add location
            args.extend(["--cidade", self.cidade_input.value])
            if self.bairro_input.value:
                endereco = f"{self.bairro_input.value}, {self.cidade_input.value}"
                args.extend(["--endereco", endereco])
            if self.tipo_select.value and self.tipo_select.value != "indiferente":
                args.extend(["--tipo_imovel", self.tipo_select.value])
            
            # Run scraper
            result = subprocess.run(
                args,
                capture_output=True,
                text=True,
                timeout=600  # 10 minute timeout
            )
            
            if result.returncode != 0:
                error_msg = result.stderr[:200] if result.stderr else f"Código de erro: {result.returncode}"
                self.update_status(f"Erro: {error_msg}", "error")
                return
            
            # Parse results
            try:
                data = json.loads(result.stdout)
            except json.JSONDecodeError:
                self.update_status("Erro ao processar resultado JSON", "error")
                return
            
            # Convert to Property objects and display
            self.properties = []
            results = data if isinstance(data, list) else data.get("results", [])
            
            for idx, item in enumerate(results):
                if not self.scraping:
                    break
                
                prop = Property(
                    id=item.get("id", f"prop-{idx}"),
                    nome=item.get("nome", f"Imóvel {idx + 1}"),
                    imagem=sanitize_image(item.get("imagem", "")),
                    valor=item.get("valor", "R$ 0"),
                    m2=item.get("m2", "0 m²"),
                    localizacao=item.get("localizacao", ""),
                    link=item.get("link", "#"),
                    quartos=item.get("quartos", ""),
                    garagem=item.get("garagem", "0"),
                    banhos=item.get("banhos", item.get("banheiros", "")),
                    site=item.get("site", "")
                )
                # Avoid duplicates
                if not any(p.link == prop.link for p in self.properties):
                    self.properties.append(prop)
                    self.display_property(prop, self.properties_view)
            
            self.refresh_properties_display()
            self.update_status(f"{len(self.properties)} imóveis encontrados", "success")
            
        except subprocess.TimeoutExpired:
            self.update_status("Timeout no scraping (10 minutos)", "error")
        except Exception as ex:
            self.update_status(f"Erro: {str(ex)[:100]}", "error")
        finally:
            self.scraping = False
            self.start_button.disabled = False
            self.stop_button.disabled = True
            self.update_stats()
            self.page.update()
    
    def display_property(self, prop: Property, view: ft.Column = None):
        """Display a single property card"""
        if view is None:
            view = self.properties_view
        
        # Property card
        card_content = ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                ft.Text(prop.nome, size=14, weight="bold", max_lines=2),
                                ft.Row(
                                    controls=[
                                        ft.Icon(ft.icons.LOCATION_ON, size=16, color="gray"),
                                        ft.Text(prop.localizacao, size=11, color="gray", max_lines=1),
                                    ],
                                    spacing=4,
                                ),
                                ft.Text(prop.valor, size=16, weight="bold", color="#16a34a"),
                            ],
                            expand=True,
                            spacing=4,
                        ),
                        ft.Column(
                            controls=[
                                ft.ElevatedButton(
                                    text="❤️",
                                    on_click=lambda e, p=prop: self.like_property(p),
                                    width=50,
                                ),
                                ft.ElevatedButton(
                                    text="👎",
                                    on_click=lambda e, p=prop: self.dislike_property(p),
                                    width=50,
                                ),
                            ],
                            spacing=4,
                        ),
                    ],
                    spacing=12,
                    vertical_alignment="start",
                ),
                ft.Row(
                    controls=[x for x in [
                        ft.Chip(label=ft.Text(prop.m2, size=10)) if prop.m2 else None,
                        ft.Chip(label=ft.Text(prop.quartos, size=10)) if prop.quartos else None,
                        ft.Chip(label=ft.Text(f"{prop.garagem} vagas", size=10)) if prop.garagem else None,
                        ft.Chip(label=ft.Text(f"{prop.banhos}", size=10)) if prop.banhos else None,
                        ft.Chip(label=ft.Text(prop.site, size=10)) if prop.site else None,
                    ] if x is not None],
                    spacing=4,
                    wrap=True,
                ),
                ft.Row(
                    controls=[
                        ft.TextButton(
                            text="Abrir link",
                            on_click=lambda e, link=prop.link: self.open_link(link),
                        ),
                        ft.TextButton(
                            text="Adicionar tag",
                            on_click=lambda e, p=prop: self.show_tag_dialog(p),
                        ),
                    ],
                    spacing=8,
                ),
                ft.Text(", ".join(prop.tags), size=10, color="blue", max_lines=2) if prop.tags else None,
            ],
            spacing=8,
        )
        
        # Filter out None values
        card_content.controls = [c for c in card_content.controls if c is not None]
        
        card = ft.Card(
            content=ft.Container(
                content=card_content,
                padding=12,
            ),
            margin=ft.margin.only(bottom=8),
        )
        
        view.controls.append(card)
        self.page.update()
    
    def like_property(self, prop: Property):
        """Mark property as liked"""
        if prop not in self.liked_properties:
            self.liked_properties.append(prop)
            self.properties = [p for p in self.properties if p.link != prop.link]
        self.show_message(f"❤️ Adicionado aos favoritos", "success")
        self.refresh_properties_display()
        self.refresh_liked_view()
        self.update_stats()
    
    def dislike_property(self, prop: Property):
        """Mark property as disliked"""
        if prop not in self.disliked_properties:
            self.disliked_properties.append(prop)
            self.properties = [p for p in self.properties if p.link != prop.link]
        self.show_message(f"👎 Rejeitado", "info")
        self.refresh_properties_display()
        self.refresh_disliked_view()
        self.update_stats()
    
    def refresh_properties_display(self):
        """Refresh the properties display with current filters and sorting"""
        self.properties_view.controls.clear()
        
        # Apply filters
        filtered = self.apply_filters(self.properties)
        
        # Apply sorting
        sorted_props = self.sort_properties(filtered)
        
        # Display
        for prop in sorted_props:
            self.display_property(prop, self.properties_view)
        
        self.page.update()
    
    def refresh_liked_view(self):
        """Refresh liked properties view"""
        self.liked_view.controls.clear()
        for prop in self.liked_properties:
            self.display_property(prop, self.liked_view)
        self.page.update()
    
    def refresh_disliked_view(self):
        """Refresh disliked properties view"""
        self.disliked_view.controls.clear()
        for prop in self.disliked_properties:
            self.display_property(prop, self.disliked_view)
        self.page.update()
    
    def clear_liked_properties(self, e):
        """Clear all liked properties"""
        if len(self.liked_properties) == 0:
            self.show_message("Nenhum imóvel favorito", "info")
            return
        self.liked_properties.clear()
        self.refresh_liked_view()
        self.update_stats()
        self.show_message("Favoritos limpos", "success")
    
    def clear_disliked_properties(self, e):
        """Clear all disliked properties"""
        if len(self.disliked_properties) == 0:
            self.show_message("Nenhum imóvel rejeitado", "info")
            return
        self.disliked_properties.clear()
        self.refresh_disliked_view()
        self.update_stats()
        self.show_message("Rejeitados limpos", "success")
    
    def refresh_ranking(self, e):
        """Refresh ranking display"""
        self.display_ranking()
    
    def display_ranking(self):
        """Display ranking view based on preferences"""
        self.ranking_view.controls.clear()
        
        if not self.liked_properties:
            empty = ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Icon(ft.Icons.FAVORITE, size=64, color="gray"),
                        ft.Text("Nenhum imóvel nos favoritos", size=16, color="gray"),
                        ft.Text("Adicione imóveis aos favoritos para gerar ranking", size=12, color="gray"),
                    ],
                    horizontal_alignment="center",
                    spacing=16,
                ),
                alignment=ft.alignment.center,
            )
            self.ranking_view.controls.append(empty)
            self.page.update()
            return
        
        # Simple ranking based on price (ascending)
        ranked = sorted(self.liked_properties, key=lambda p: parse_int(p.valor))
        
        for idx, prop in enumerate(ranked, 1):
            # Display with ranking number
            ranking_card = ft.Column(
                controls=[
                    ft.Row(
                        controls=[
                            ft.Text(f"#{idx}", size=20, weight="bold", color="#2563eb", width=50),
                            ft.Text(f"⭐ Score: {max(0, 100 - (idx * 5))}%", size=14, color="orange", width=150),
                        ],
                        spacing=8,
                    ),
                    self.create_property_display(prop),
                ],
                spacing=8,
            )
            card = ft.Card(
                content=ft.Container(
                    content=ranking_card,
                    padding=12,
                ),
                margin=ft.margin.only(bottom=8),
            )
            self.ranking_view.controls.append(card)
        
        self.page.update()
    
    def create_property_display(self, prop: Property) -> ft.Column:
        """Create a property display without like/dislike buttons"""
        return ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                ft.Text(prop.nome, size=14, weight="bold", max_lines=2),
                                ft.Row(
                                    controls=[
                                        ft.Icon(ft.icons.LOCATION_ON, size=16, color="gray"),
                                        ft.Text(prop.localizacao, size=11, color="gray", max_lines=1),
                                    ],
                                    spacing=4,
                                ),
                                ft.Text(prop.valor, size=16, weight="bold", color="#16a34a"),
                            ],
                            expand=True,
                            spacing=4,
                        ),
                        ft.TextButton(
                            text="Abrir",
                            on_click=lambda e, link=prop.link: self.open_link(link),
                        ),
                    ],
                    spacing=12,
                ),
                ft.Row(
                    controls=[x for x in [
                        ft.Chip(label=ft.Text(prop.m2, size=10)) if prop.m2 else None,
                        ft.Chip(label=ft.Text(prop.quartos, size=10)) if prop.quartos else None,
                        ft.Chip(label=ft.Text(f"{prop.garagem} vagas", size=10)) if prop.garagem else None,
                        ft.Chip(label=ft.Text(f"{prop.banhos}", size=10)) if prop.banhos else None,
                        ft.Chip(label=ft.Text(prop.site, size=10)) if prop.site else None,
                    ] if x is not None],
                    spacing=4,
                    wrap=True,
                ),
                ft.Text(", ".join(prop.tags), size=10, color="blue", max_lines=2) if prop.tags else None,
            ],
            spacing=8,
        )
    
    def apply_filters(self, properties: List[Property]) -> List[Property]:
        """Apply all active filters to properties"""
        filtered = []
        
        for prop in properties:
            # Price filter
            valor_num = parse_int(prop.valor)
            valor_min = parse_int(self.valormin_input.value) if self.valormin_input.value else 0
            valor_max = parse_int(self.valormax_input.value) if self.valormax_input.value else float('inf')
            
            if valor_num < valor_min or valor_num > valor_max:
                continue
            
            # Size filter
            area_num = parse_int(prop.m2)
            area_min = parse_int(self.area_min_input.value) if self.area_min_input.value else 0
            area_max = parse_int(self.area_max_input.value) if self.area_max_input.value else float('inf')
            
            if area_num < area_min or area_num > area_max:
                continue
            
            # Rooms filter
            if self.quartos_input.value:
                allowed_quartos = [x.strip() for x in self.quartos_input.value.split(',')]
                prop_quartos = parse_int(prop.quartos)
                matched = False
                for q in allowed_quartos:
                    if q.endswith('+'):
                        if prop_quartos >= parse_int(q):
                            matched = True
                            break
                    else:
                        if prop_quartos == parse_int(q):
                            matched = True
                            break
                if not matched:
                    continue
            
            # Parking filter
            if self.vagas_input.value:
                allowed_vagas = [x.strip() for x in self.vagas_input.value.split(',')]
                prop_vagas = parse_int(prop.garagem)
                matched = False
                for v in allowed_vagas:
                    if v.endswith('+'):
                        if prop_vagas >= parse_int(v):
                            matched = True
                            break
                    else:
                        if prop_vagas == parse_int(v):
                            matched = True
                            break
                if not matched:
                    continue
            
            # Bathrooms filter
            if self.banhos_input.value:
                allowed_banhos = [x.strip() for x in self.banhos_input.value.split(',')]
                prop_banhos = parse_int(prop.banhos)
                matched = False
                for b in allowed_banhos:
                    if b.endswith('+'):
                        if prop_banhos >= parse_int(b):
                            matched = True
                            break
                    else:
                        if prop_banhos == parse_int(b):
                            matched = True
                            break
                if not matched:
                    continue
            
            filtered.append(prop)
        
        return filtered
    
    def sort_properties(self, properties: List[Property]) -> List[Property]:
        """Sort properties based on current sort option"""
        if not self.sort_option:
            return properties
        
        field = self.sort_option.get("field", "valor")
        direction = self.sort_option.get("direction", "asc")
        reverse = direction == "desc"
        
        if field == "valor":
            return sorted(properties, key=lambda p: parse_int(p.valor), reverse=reverse)
        elif field == "tamanho":
            return sorted(properties, key=lambda p: parse_int(p.m2), reverse=reverse)
        elif field == "distancia" and self.user_location:
            return sorted(properties, key=lambda p: self.calculate_distance(p), reverse=reverse)
        else:
            return properties
    
    def calculate_distance(self, prop: Property) -> float:
        """Calculate distance from user location to property"""
        if not self.user_location or not prop.latitude or not prop.longitude:
            return float('inf')
        
        # Haversine formula
        R = 6371  # Earth's radius in km
        lat1, lon1 = self.user_location.get("latitude", 0), self.user_location.get("longitude", 0)
        lat2, lon2 = prop.latitude, prop.longitude
        
        dLat = math.radians(lat2 - lat1)
        dLon = math.radians(lon2 - lon1)
        a = math.sin(dLat/2) * math.sin(dLat/2) + \
            math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * \
            math.sin(dLon/2) * math.sin(dLon/2)
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
        return R * c
    
    def show_tag_dialog(self, prop: Property):
        """Show dialog to add tags to a property"""
        tag_input = ft.TextField(label="Nova tag", width=300)
        
        def add_tag(e):
            tag = tag_input.value.strip()
            if tag:
                if tag not in prop.tags:
                    prop.tags.append(tag)
                    self.show_message(f"Tag '{tag}' adicionada", "success")
                tag_input.value = ""
                self.refresh_liked_view()
                dlg.open = False
                self.page.update()
        
        dlg = ft.AlertDialog(
            title=ft.Text("Adicionar tag"),
            content=ft.Column(
                controls=[
                    ft.Text(f"Imóvel: {prop.nome}", size=12),
                    tag_input,
                    ft.Text("Tags atuais: " + ", ".join(prop.tags) if prop.tags else "Nenhuma tag", size=10, color="gray"),
                ],
                tight=True,
                spacing=8,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg, 'open', False), self.page.update())),
                ft.TextButton("Adicionar", on_click=add_tag),
            ],
        )
        
        self.page.dialog = dlg
        dlg.open = True
        self.page.update()
    
    def show_settings_dialog(self):
        """Show settings dialog for user preferences"""
        location_input = ft.TextField(
            label="Seu endereço",
            value=self.user_location.get("address", "") if self.user_location else "Belo Horizonte",
            width=300,
        )
        
        tamanho_priority = ft.Dropdown(
            label="Prioridade - Tamanho",
            options=[ft.dropdown.Option(str(i), f"Posição {i}") for i in range(1, 5)],
            value=str(self.pref_tamanho_priority),
            width=150,
        )
        
        tamanho_value = ft.TextField(
            label="Tamanho preferido (m²)",
            value=str(self.pref_tamanho_value),
            width=150,
        )
        
        quartos_priority = ft.Dropdown(
            label="Prioridade - Quartos",
            options=[ft.dropdown.Option(str(i), f"Posição {i}") for i in range(1, 5)],
            value=str(self.pref_quartos_priority),
            width=150,
        )
        
        quartos_value = ft.TextField(
            label="Quartos preferidos",
            value=str(self.pref_quartos_value),
            width=150,
        )
        
        def save_settings(e):
            try:
                self.pref_tamanho_priority = int(tamanho_priority.value)
                self.pref_quartos_priority = int(quartos_priority.value)
                self.pref_tamanho_value = int(tamanho_value.value)
                self.pref_quartos_value = int(quartos_value.value)
                
                if location_input.value:
                    self.user_location = {
                        "address": location_input.value,
                        "latitude": -19.9191,  # Default BH coordinates
                        "longitude": -43.9386,
                    }
                
                self.show_message("Configurações salvas", "success")
                dlg.open = False
                self.page.update()
            except Exception as ex:
                self.show_message(f"Erro ao salvar: {str(ex)}", "error")
        
        dlg = ft.AlertDialog(
            title=ft.Text("Configurações"),
            content=ft.Column(
                controls=[
                    ft.Text("Localização", weight="bold"),
                    location_input,
                    ft.Divider(),
                    ft.Text("Preferências de Ranking", weight="bold"),
                    ft.Row([tamanho_priority, tamanho_value], spacing=8),
                    ft.Row([quartos_priority, quartos_value], spacing=8),
                ],
                tight=True,
                spacing=12,
                scroll="auto",
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg, 'open', False), self.page.update())),
                ft.TextButton("Salvar", on_click=save_settings),
            ],
        )
        
        self.page.dialog = dlg
        dlg.open = True
        self.page.update()
    
    def update_stats(self):
        """Update status statistics"""
        self.status_stats.value = f"Total: {len(self.properties)} | Favoritos: {len(self.liked_properties)} | Rejeitados: {len(self.disliked_properties)}"
        self.page.update()
    
    def open_link(self, link: str):
        """Open property link in browser"""
        import webbrowser
        if link and link != "#":
            try:
                webbrowser.open(link)
            except Exception as ex:
                self.show_message(f"Erro ao abrir link: {str(ex)}", "error")
        else:
            self.show_message("Link indisponível", "error")
    
    def show_message(self, message: str, msg_type: str = "info"):
        """Show a message to the user"""
        colors = {
            "success": "#16a34a",
            "error": "#dc2626",
            "warning": "#ea580c",
            "info": "#2563eb",
        }
        
        snack = ft.SnackBar(
            ft.Text(message, color="white"),
            bgcolor=colors.get(msg_type, "#2563eb"),
            duration=3000,
        )
        self.page.overlay.append(snack)
        snack.open = True
        self.page.update()
    
    def update_status(self, message: str, status_type: str = "info"):
        """Update the status indicator and message"""
        colors = {
            "success": "#16a34a",
            "error": "#dc2626",
            "warning": "#ea580c",
            "info": "#2563eb",
        }
        
        self.status_text.value = message
        self.status_text.color = colors.get(status_type, "gray")
        self.status_indicator.bgcolor = colors.get(status_type, "gray")
        self.page.update()


def main(page: ft.Page):
    """Main entry point"""
    app = ScraperGUI(page)


if __name__ == "__main__":
    ft.app(target=main)
